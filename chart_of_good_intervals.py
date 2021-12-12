import openpyxl
from typing import List, Optional

import fuzzy.parse as parse
from fuzzy.classes import GoodInterval


def save_intervals_to_spreadsheet(filename: str, intervals: List[GoodInterval]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    init_time: int = 0
    if intervals:
        init_time = intervals[0].start
    sheet.append([f"init time={init_time}", "grade"])
    for interval in intervals:
        time_diff: int = interval.start - init_time
        sheet.append([f"{time_diff//60}:{time_diff%60:02d}", interval.grade])
    sheet.column_dimensions["A"].width = 20  # col A are the times
    workbook.save(filename=filename)


def draw_chart_of_intervals_on_spreadsheet(filename: str) -> None:
    """Draws the bar chart of intervals read from the spreadsheet and save it
    back to the spreadsheet.

    Arguments:
        filename: The spreadsheet which contains the intervals.
    """
    workbook = openpyxl.load_workbook(filename=filename)
    sheet = workbook.active

    chart = openpyxl.chart.BarChart()
    chart.title = "Bar Chart"
    chart.y_axis.title = "Concentration"
    chart.x_axis.title = "Time"

    data = openpyxl.chart.Reference(sheet,
                                    min_col=2, max_col=2,
                                    min_row=1, max_row=sheet.max_row)
    cats = openpyxl.chart.Reference(sheet,
                                    min_col=1, max_col=1,
                                    min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.width = 20
    sheet.add_chart(chart, "F3")

    workbook.save(filename=filename)


if __name__ == "__main__":
    spreadsheet: str = "good_intervals.xlsx"

    intervals: List[GoodInterval] = parse.read_intervals_from_json("good_intervals.json")
    save_intervals_to_spreadsheet(spreadsheet, intervals)
    draw_chart_of_intervals_on_spreadsheet(spreadsheet)
